# giyul_logic.py
import io
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Alignment

# ---------- твои функции ----------

def parse_amount(val):
    if val is None or val == "":
        raise ValueError("empty")
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s == "":
        raise ValueError("empty")
    s = s.replace(",", "")
    return float(s)


def detect_headers(ws):
    candidates = [1, 2]
    chosen_row = None
    headers = {}

    for row_idx in candidates:
        row_cells = ws[row_idx]
        row_values = [str(c.value).strip() if c.value is not None else "" for c in row_cells]
        if any(v for v in row_values):
            tmp_headers = {str(c.value).strip(): c.column for c in ws[row_idx] if c.value}
            if "חשבון" in tmp_headers and "חוב לחשבונית" in tmp_headers:
                chosen_row = row_idx
                headers = tmp_headers
                break
            if not headers:
                chosen_row = row_idx
                headers = tmp_headers

    if chosen_row is None:
        chosen_row = 1
        headers = {str(c.value).strip(): c.column for c in ws[1] if c.value}

    return chosen_row, headers


GREEN_RGB = "FF00FF00"
ORANGE_RGB = "FFFFA500"
PURPLE_RGB = "FFCC99FF"
BLUE_RGB = "FFADD8E6"

GREEN_FILL = PatternFill(start_color=GREEN_RGB, end_color=GREEN_RGB, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=ORANGE_RGB, end_color=ORANGE_RGB, fill_type="solid")
PURPLE_FILL = PatternFill(start_color=PURPLE_RGB, end_color=PURPLE_RGB, fill_type="solid")
BLUE_FILL = PatternFill(start_color=BLUE_RGB, end_color=BLUE_RGB, fill_type="solid")


def cell_rgb(cell):
    try:
        return cell.fill.start_color.rgb
    except Exception:
        return None


def has_any_color(cell):
    return cell.fill.fill_type == "solid" and cell_rgb(cell) in {
        GREEN_RGB,
        ORANGE_RGB,
        PURPLE_RGB,
        BLUE_RGB,
    }


def ensure_summary_sheet(wb, title, counts):
    if title in wb.sheetnames:
        ws_sum = wb[title]
        for row in ws_sum.iter_rows():
            for c in row:
                c.value = None
    else:
        ws_sum = wb.create_sheet(title)

    ws_sum["A1"] = "מס ספק"
    ws_sum["B1"] = "כמות שורות מותאמות"

    r = 2
    for acc, cnt in counts.items():
        if acc is None or cnt <= 0:
            continue
        ws_sum.cell(r, 1, acc)
        ws_sum.cell(r, 2, cnt)
        r += 1


def process_workbook(wb):
    ws = wb.active

    header_row, headers = detect_headers(ws)

    col_acc = headers.get("חשבון")
    col_amt = headers.get("חוב לחשבונית")
    col_type = headers.get("סוג תנועה")
    col_name = headers.get("תאור חשבון") or headers.get("שם ספק") or headers.get("תיאור חשבון")
    col_pay = headers.get("תאריך תשלום")

    if col_acc is None or col_amt is None:
        raise ValueError("לא נמצאו עמודות 'חשבון' ו/או 'חוב לחשבונית'. ודאי ששמות הכותרות כתובים בדיוק כך.")

    if col_name is None:
        col_name = 3
    if col_pay is None:
        col_pay = 4

    data_start_row = header_row + 1

    # ===== логика 1 – зелёный 100% =====
    groups = defaultdict(list)
    for row in ws.iter_rows(min_row=data_start_row):
        acc = row[col_acc - 1].value
        groups[acc].append(row)

    green_counts = defaultdict(int)

    for acc, rows in groups.items():
        pos, neg = [], []
        for r in rows:
            cell = r[col_amt - 1]
            try:
                v = parse_amount(cell.value)
            except Exception:
                continue
            if v > 0:
                pos.append((v, r))
            elif v < 0:
                neg.append((v, r))

        used_neg = set()
        for pval, prow in pos:
            for ni, (nval, nrow) in enumerate(neg):
                if ni in used_neg:
                    continue
                if abs(pval + nval) < 1e-6:
                    prow[col_amt - 1].fill = GREEN_FILL
                    nrow[col_amt - 1].fill = GREEN_FILL
                    green_counts[acc] += 2
                    used_neg.add(ni)
                    break

    ensure_summary_sheet(wb, "התאמה 100%", green_counts)

    # ===== логика 3 – оранжевый 80% =====
    orange_counts = defaultdict(int)

    for acc, rows in groups.items():
        pos, neg = [], []
        for r in rows:
            cell = r[col_amt - 1]
            if has_any_color(cell):
                continue
            try:
                v = parse_amount(cell.value)
            except Exception:
                continue
            if v > 0:
                pos.append((v, r))
            elif v < 0:
                neg.append((v, r))

        used_neg = set()
        for pval, prow in pos:
            pc = prow[col_amt - 1]
            if has_any_color(pc):
                continue
            for ni, (nval, nrow) in enumerate(neg):
                if ni in used_neg:
                    continue
                nc = nrow[col_amt - 1]
                if has_any_color(nc):
                    continue
                if abs(pval + nval) <= 2:
                    pc.fill = ORANGE_FILL
                    nc.fill = ORANGE_FILL
                    orange_counts[acc] += 2
                    used_neg.add(ni)
                    break

    ensure_summary_sheet(wb, "התאמה 80%", orange_counts)

    # ===== логика 5 – фиолетовый глобальный =====
    purple_counts = defaultdict(int)
    eligible = []

    for row in ws.iter_rows(min_row=data_start_row):
        cell = row[col_amt - 1]
        if has_any_color(cell):
            continue
        try:
            v = parse_amount(cell.value)
        except Exception:
            continue
        if v == 0:
            continue
        acc = row[col_acc - 1].value
        eligible.append((v, acc, row))

    pos = [x for x in eligible if x[0] > 0]
    neg = [x for x in eligible if x[0] < 0]

    used_pos, used_neg = set(), set()

    for pi, (pval, pacc, prow) in enumerate(pos):
        if pi in used_pos:
            continue
        pc = prow[col_amt - 1]
        if has_any_color(pc):
            continue
        for ni, (nval, nacc, nrow) in enumerate(neg):
            if ni in used_neg:
                continue
            nc = nrow[col_amt - 1]
            if has_any_color(nc):
                continue
            if abs(pval + nval) <= 2:
                pc.fill = PURPLE_FILL
                nc.fill = PURPLE_FILL
                used_pos.add(pi)
                used_neg.add(ni)
                purple_counts[pacc] += 1
                purple_counts[nacc] += 1
                break

    ensure_summary_sheet(wb, "בדיקת ספקים", purple_counts)

    # ===== логика 6 – голубой "העב" =====
    rows_mail = []

    for row in ws.iter_rows(min_row=data_start_row):
        if col_type is None:
            continue
        tval = row[col_type - 1].value
        tval = str(tval).strip() if tval is not None else ""
        cell = row[col_amt - 1]
        if tval == "העב" and not has_any_color(cell):
            cell.fill = BLUE_FILL
            rows_mail.append(
                (
                    row[col_name - 1].value,
                    row[col_pay - 1].value,
                    row[col_amt - 1].value,
                )
            )

    # ===== логика 7 – "מיילים לספק" =====
    if "מיילים לספק" in wb.sheetnames:
        ws_mail = wb["מיילים לספק"]
        for r in ws_mail.iter_rows():
            for c in r:
                c.value = None
    else:
        ws_mail = wb.create_sheet("מיילים לספק")

    ws_mail["A1"] = "שם ספק"
    ws_mail["B1"] = "תאריך תשלום"
    ws_mail["C1"] = "חוב לחשבונית"
    ws_mail["D1"] = "טקסט מייל"

    company_name = ws["C1"].value if ws["C1"].value is not None else ""

    row_idx = 2
    for name, pay, debt in rows_mail:
        ws_mail.cell(row_idx, 1, name)

        if isinstance(pay, datetime):
            date_str = pay.strftime("%d/%m/%y")
        else:
            date_str = str(pay) if pay is not None else ""
        ws_mail.cell(row_idx, 2, date_str)

        try:
            amount = abs(parse_amount(debt))
        except Exception:
            amount = debt
        ws_mail.cell(row_idx, 3, amount)

        msg = (
            f"שלום ל-{name}\n"
            f"חסרה לנו חשבונית עבור תשלום:\n"
            f"תאריך - {date_str}\n"
            f"על סכום - {amount}\n"
            f"בתודה מראש,\n"
            f"הנהלת חשבונות של {company_name}"
        )
        cell_msg = ws_mail.cell(row_idx, 4, msg)
        cell_msg.alignment = Alignment(wrap_text=True)

        row_idx += 1

    for sh in wb.worksheets:
        sh.sheet_view.rightToLeft = True

    return wb
